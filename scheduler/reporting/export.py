"""Excel export helpers for solved schedules."""

from __future__ import annotations

import importlib.util
from pathlib import Path
from typing import Dict, List, Tuple
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from ortools.sat.python import cp_model

from scheduler.config import FRONT_DESK_ROLE, SLOT_MINUTES, slots_to_hours
from scheduler.reporting.stats import aggregate_department_hours


def export_schedule_to_excel(
    status,
    solver,
    employees,
    days,
    time_slots,
    slot_names,
    qual,
    work,
    assign,
    weekly_hour_limits,
    target_weekly_hours,
    roles,
    department_roles,
    role_display_names,
    department_hour_targets,
    department_max_hours,
    output_path: Path,
    primary_frontdesk_department,
):
    """Export the generated schedule to an Excel workbook with formatted sheets."""
    if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        return

    role_columns = [FRONT_DESK_ROLE] + department_roles
    frontdesk_comment_for = _build_frontdesk_comment_lookup(
        employees, qual, primary_frontdesk_department, role_display_names
    )

    daily_tables = []
    weekly_rows = []
    role_headers = [role_display_names[role] for role in role_columns]
    weekly_columns = ["Day", "Time"] + role_headers

    for day in days:
        day_rows = []
        for t in time_slots:
            cell_values = []
            for role in role_columns:
                workers = [
                    e for e in employees if (e, day, t, role) in assign and solver.value(assign[(e, day, t, role)])
                ]
                cell_values.append(", ".join(workers) if workers else ("UNCOVERED" if role == FRONT_DESK_ROLE else ""))
            day_rows.append([slot_names[t], *cell_values])
            weekly_rows.append([day, slot_names[t], *cell_values])
        daily_tables.append((f"{day} Schedule", ["Time"] + role_headers, day_rows))

    summary_rows = []
    for e in employees:
        total_slots = 0
        days_worked = []
        for d in days:
            day_slots = sum(solver.value(work[e, d, t]) for t in time_slots)
            if day_slots > 0:
                total_slots += day_slots
                days_worked.append(f"{d}({slots_to_hours(day_slots):.1f}h)")
        target_hours = target_weekly_hours.get(e, 0)
        max_hours = weekly_hour_limits.get(e, 0)
        total_hours = slots_to_hours(total_slots)
        hit_target = abs(total_hours - target_hours) <= (SLOT_MINUTES / 60)
        summary_rows.append(
            [
                e,
                ", ".join(sorted(qual[e])),
                total_hours,
                target_hours,
                max_hours,
                "✓" if hit_target else "",
                ", ".join(days_worked) if days_worked else "None",
            ]
        )
    summary_columns = [
        "Employee",
        "Qualifications",
        "Hours Worked",
        "Target Hours",
        "Max Hours",
        "Hit Target",
        "Days Worked",
    ]

    distribution_rows = []
    role_totals = {role: 0 for role in roles}
    for d in days:
        row = [d]
        for role in roles:
            slot_count = sum(
                solver.value(assign[(e, d, t, role)]) if (e, d, t, role) in assign else 0 for e in employees for t in time_slots
            )
            role_totals[role] += slot_count
            row.append(slots_to_hours(slot_count))
        distribution_rows.append(row)
    total_row = ["TOTAL"] + [slots_to_hours(role_totals[r]) for r in roles]
    distribution_rows.append(total_row)
    distribution_columns = ["Day"] + [role_display_names[role] for role in roles]

    _, _, department_breakdown = aggregate_department_hours(
        solver, employees, days, time_slots, assign, department_roles, qual, primary_frontdesk_department
    )

    dept_summary_headers = [
        "Department",
        "Actual Hours",
        "Target Hours",
        "Max Hours",
        "Delta (Actual-Target)",
        "Focused Hours",
        "Dual Hours Total",
        "Dual Hours Counted",
    ]
    dept_summary_rows = []
    for role in department_roles:
        stats = department_breakdown[role]
        actual_hours = stats["actual_hours"]
        target = department_hour_targets.get(role)
        max_hours = department_max_hours.get(role)
        delta = actual_hours - target if target is not None else ""
        dept_summary_rows.append(
            [
                role_display_names[role],
                actual_hours,
                target if target is not None else "",
                max_hours if max_hours is not None else "",
                delta,
                stats["focused_hours"],
                stats["dual_hours_total"],
                stats["dual_hours_counted"],
            ]
        )

    engine = None
    for candidate in ("xlsxwriter", "openpyxl"):
        if importlib.util.find_spec(candidate):
            engine = candidate
            break

    if engine is None:
        _write_minimal_xlsx(output_path, weekly_columns, weekly_rows)
        return

    with pd.ExcelWriter(output_path, engine=engine) as writer:
        df_weekly = pd.DataFrame(weekly_rows, columns=weekly_columns)
        df_weekly.to_excel(writer, sheet_name="Weekly Grid", index=False)
        _autosize_columns(writer, "Weekly Grid", df_weekly)
        _add_frontdesk_comments_table(
            writer=writer,
            engine=engine,
            sheet_name="Weekly Grid",
            rows=weekly_rows,
            value_column_idx=2,  # front desk column
            row_offset=1,
            comment_lookup=frontdesk_comment_for,
        )

        for sheet_name, columns, rows in daily_tables:
            df_day = pd.DataFrame(rows, columns=columns)
            df_day.to_excel(writer, sheet_name=sheet_name, index=False)
            _autosize_columns(writer, sheet_name, df_day)
            _add_frontdesk_comments_table(
                writer=writer,
                engine=engine,
                sheet_name=sheet_name,
                rows=rows,
                value_column_idx=1,  # front desk column
                row_offset=1,
                comment_lookup=frontdesk_comment_for,
            )

        df_summary = pd.DataFrame(summary_rows, columns=summary_columns)
        df_summary.to_excel(writer, sheet_name="Employee Summary", index=False)
        _autosize_columns(writer, "Employee Summary", df_summary)

        df_distribution = pd.DataFrame(distribution_rows, columns=distribution_columns)
        df_distribution.to_excel(writer, sheet_name="Role Distribution", index=False)
        _autosize_columns(writer, "Role Distribution", df_distribution)
        if dept_summary_rows:
            df_dept = pd.DataFrame(dept_summary_rows, columns=dept_summary_headers)
            df_dept.to_excel(writer, sheet_name="Department Summary", index=False)
            _autosize_columns(writer, "Department Summary", df_dept)


def _format_time_range(start_str: str, end_str: str) -> str:
    def to_minutes(s: str) -> int:
        h, m = map(int, s.split(":"))
        return h * 60 + m

    def fmt(minutes: int) -> str:
        h = (minutes // 60) % 24
        m = minutes % 60
        suffix = "AM" if h < 12 else "PM"
        hour12 = h % 12 or 12
        return f"{hour12}{':' + str(m).zfill(2) if m else ''}{suffix}"

    start_min = to_minutes(start_str)
    end_min = to_minutes(end_str)
    # end_str represents the start of the final slot; add one slot for the true end
    end_min += SLOT_MINUTES
    return f"{fmt(start_min)}-{fmt(end_min)}"


def _collect_intervals(assign, solver, employees, days, T, time_slots, role):
    intervals: Dict[str, List[Tuple[str, int, int]]] = {day: [] for day in days}
    for day in days:
        for e in employees:
            slots = [t for t in T if (e, day, t, role) in assign and solver.value(assign[(e, day, t, role)])]
            if not slots:
                continue
            slots.sort()
            start = prev = slots[0]
            for s in slots[1:] + [None]:
                if s is not None and s == prev + 1:
                    prev = s
                    continue
                intervals[day].append((e, start, prev))
                if s is not None:
                    start = prev = s
        # sort by start time then name for stable ordering
        intervals[day].sort(key=lambda x: (x[1], x[0]))
    return intervals


def _find_coverage_gaps(intervals: Dict[str, List[Tuple[str, int, int]]], days, T) -> Dict[str, List[Tuple[str, int, int]]]:
    """Find time slots with no coverage and return them as UNCOVERED intervals."""
    gaps: Dict[str, List[Tuple[str, int, int]]] = {day: [] for day in days}
    
    for day in days:
        # Create a set of all covered slots for this day
        covered_slots = set()
        for _, start, end in intervals.get(day, []):
            for t in range(start, end + 1):
                covered_slots.add(t)
        
        # Find uncovered slots
        uncovered_slots = [t for t in T if t not in covered_slots]
        if not uncovered_slots:
            continue
        
        # Group consecutive uncovered slots into intervals
        uncovered_slots.sort()
        start = prev = uncovered_slots[0]
        for s in uncovered_slots[1:] + [None]:
            if s is not None and s == prev + 1:
                prev = s
                continue
            gaps[day].append(("UNCOVERED", start, prev))
            if s is not None:
                start = prev = s
    
    return gaps


def _merge_intervals_with_gaps(
    intervals: Dict[str, List[Tuple[str, int, int]]], 
    gaps: Dict[str, List[Tuple[str, int, int]]], 
    days
) -> Dict[str, List[Tuple[str, int, int, bool]]]:
    """Merge coverage intervals with gap intervals, sorted by start time.
    
    Returns intervals with a 4th element indicating if it's an UNCOVERED gap.
    """
    merged: Dict[str, List[Tuple[str, int, int, bool]]] = {day: [] for day in days}
    
    for day in days:
        # Add regular intervals (not gaps)
        for name, start, end in intervals.get(day, []):
            merged[day].append((name, start, end, False))
        
        # Add gap intervals
        for name, start, end in gaps.get(day, []):
            merged[day].append((name, start, end, True))
        
        # Sort by start time
        merged[day].sort(key=lambda x: (x[1], x[0]))
    
    return merged


def _build_frontdesk_comment_lookup(employees, qual, primary_frontdesk_department, role_display_names):
    """Return a callable mapping employee name to an optional front desk comment."""
    multi_dept_employees = set(
        e for e in employees if e in qual and len([r for r in qual[e] if r != FRONT_DESK_ROLE]) > 1
    )

    def lookup(name: str) -> str | None:
        if name not in multi_dept_employees:
            return None
        primary = primary_frontdesk_department.get(name)
        if not primary:
            return None
        display = role_display_names.get(primary, primary.replace("_", " ").title())
        return f"Half-time counts toward {display}"

    return lookup


def _add_frontdesk_comments_table(writer, engine, sheet_name, rows, value_column_idx, row_offset, comment_lookup):
    """Attach comments to front desk cells for multi-department employees."""
    if engine not in ("xlsxwriter", "openpyxl"):
        return
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None:
        return

    for idx, row in enumerate(rows):
        if value_column_idx >= len(row):
            continue
        cell_value = row[value_column_idx]
        if not isinstance(cell_value, str):
            continue
        cell_value = cell_value.strip()
        if not cell_value or cell_value.upper() == "UNCOVERED":
            continue
        # Front desk should only have one worker; if multiple, skip to avoid ambiguity
        if "," in cell_value:
            continue
        comment = comment_lookup(cell_value)
        if not comment:
            continue
        if engine == "xlsxwriter":
            worksheet.write_comment(row_offset + idx, value_column_idx, comment)
        elif engine == "openpyxl":
            try:
                from openpyxl.comments import Comment
            except ImportError:
                continue
            excel_row = row_offset + idx + 1  # openpyxl is 1-indexed and header row is 1
            excel_col = value_column_idx + 1
            cell = worksheet.cell(row=excel_row, column=excel_col)
            cell.comment = Comment(comment, "scheduler")


def export_formatted_schedule(
    status,
    solver,
    employees,
    days,
    T,
    time_slot_starts,
    slot_names,
    qual,
    assign,
    department_roles,
    role_display_names,
    department_hour_targets,
    department_max_hours,
    primary_frontdesk_department,
    output_path: Path,
):
    """Create an alternate, styled schedule file with per-department day grids."""
    if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        return

    try:
        import xlsxwriter  # noqa: F401
    except ImportError:
        return  # skip formatted export if xlsxwriter is unavailable

    formatted_path = output_path.with_name(f"{output_path.stem}-formatted{output_path.suffix}")

    role_direct_slots, _, department_breakdown = aggregate_department_hours(
        solver, employees, days, T, assign, department_roles, qual, primary_frontdesk_department
    )

    # Build department order: front desk then department roles
    ordered_roles = [FRONT_DESK_ROLE] + department_roles
    frontdesk_comment_for = _build_frontdesk_comment_lookup(
        employees, qual, primary_frontdesk_department, role_display_names
    )

    # Gather intervals per role/day
    intervals_by_role = {
        role: _collect_intervals(assign, solver, employees, days, T, time_slot_starts, role)
        for role in ordered_roles
    }

    # Full day names mapping
    day_full_names = {
        "Mon": "Monday",
        "Tue": "Tuesday",
        "Wed": "Wednesday",
        "Thu": "Thursday",
        "Fri": "Friday",
    }

    workbook = xlsxwriter.Workbook(formatted_path)
    ws = workbook.add_worksheet("Schedule")

    # Title format - large Calibri font
    title_fmt = workbook.add_format({
        "font_name": "Calibri",
        "font_size": 26,
        "bold": False,
    })

    # Header format - cyan background, bold Calibri, all borders, LEFT aligned
    header_fmt = workbook.add_format({
        "bold": True,
        "font_name": "Calibri",
        "font_color": "black",
        "bg_color": "#00AEEF",
        "align": "left",
        "valign": "vcenter",
        "border": 1,
    })

    # Bold format for department names - Calibri Bold
    bold_fmt = workbook.add_format({
        "bold": True,
        "font_name": "Calibri",
    })

    # Text format for stats - Calibri regular
    stats_fmt = workbook.add_format({
        "font_name": "Calibri",
        "align": "left",
        "valign": "top",
    })

    # Grid cell format - Calibri with borders
    cell_fmt = workbook.add_format({
        "font_name": "Calibri",
        "align": "left",
        "valign": "top",
        "border": 1,
    })

    # Empty cell format - borders only for grid structure
    empty_cell_fmt = workbook.add_format({
        "border": 1,
    })

    # Uncovered cell format - red text for front desk gaps
    uncovered_fmt = workbook.add_format({
        "font_name": "Calibri",
        "font_color": "#FF0000",
        "align": "left",
        "valign": "top",
        "border": 1,
        "bold": True,
    })

    # Column setup: A stats, then pairs per day
    ws.set_column("A:A", 28)
    for idx in range(5):  # 5 days
        name_col = 1 + idx * 2
        time_col = 2 + idx * 2
        ws.set_column(name_col, name_col, 14)
        ws.set_column(time_col, time_col, 18)

    # Write title at top
    ws.write(0, 0, "Title", title_fmt)
    row = 1  # Start content on row 1 (right after title)

    for role in ordered_roles:
        intervals = intervals_by_role[role]
        max_len = max((len(v) for v in intervals.values()), default=0)
        display = role_display_names.get(role, role.replace("_", " ").title())

        if role == FRONT_DESK_ROLE:
            # FD: Row has FD label + day headers on same row
            counted = slots_to_hours(role_direct_slots[role])
            ws.write(row, 0, f"FD: {counted:.1f}", bold_fmt)
            
            # Day headers on same row as FD
            for idx, day in enumerate(days):
                full_day = day_full_names.get(day, day)
                name_col = 1 + idx * 2
                ws.write(row, name_col, full_day, header_fmt)
                ws.write(row, name_col + 1, "", header_fmt)
            row += 1
            
            # Find coverage gaps and merge with intervals
            gaps = _find_coverage_gaps(intervals, days, T)
            merged_intervals = _merge_intervals_with_gaps(intervals, gaps, days)
            merged_max_len = max((len(v) for v in merged_intervals.values()), default=0)
            
            # Schedule data rows - minimum 3 rows with borders
            grid_rows = max(3, merged_max_len)
            for i in range(grid_rows):
                for idx, day in enumerate(days):
                    entries = merged_intervals.get(day, [])
                    name_col = 1 + idx * 2
                    time_col = 1 + idx * 2 + 1
                    
                    if i < len(entries):
                        name, start, end, is_uncovered = entries[i]
                        start_str = time_slot_starts[start]
                        end_str = time_slot_starts[end]
                        
                        if is_uncovered:
                            # Use red text for uncovered slots
                            ws.write(row + i, name_col, name, uncovered_fmt)
                            ws.write(row + i, time_col, _format_time_range(start_str, end_str), cell_fmt)
                        else:
                            ws.write(row + i, name_col, name, cell_fmt)
                            ws.write(row + i, time_col, _format_time_range(start_str, end_str), cell_fmt)
                            comment = frontdesk_comment_for(name)
                            if comment:
                                ws.write_comment(row + i, name_col, comment)
                    else:
                        # Empty cell with border
                        ws.write(row + i, name_col, "", empty_cell_fmt)
                        ws.write(row + i, time_col, "", empty_cell_fmt)
            row += grid_rows
        else:
            # Department: Row 1 has dept name + day headers
            stats = department_breakdown[role]
            counted = stats["focused_hours"] + stats["dual_hours_counted"]
            actual = stats["focused_hours"] + stats["dual_hours_total"]
            dual_counted = stats["dual_hours_counted"]
            dual_actual = stats["dual_hours_total"]
            focused = stats["focused_hours"]
            
            # Dept name + day headers on same row
            ws.write(row, 0, f"{display}: {counted:.1f} ({actual:.1f})", bold_fmt)
            for idx, day in enumerate(days):
                full_day = day_full_names.get(day, day)
                name_col = 1 + idx * 2
                ws.write(row, name_col, full_day, header_fmt)
                ws.write(row, name_col + 1, "", header_fmt)
            row += 1
            
            # Grid rows: Dual line + row 0 data, Focused line + row 1 data, then more rows
            # Minimum 3 rows of bordered cells
            grid_rows = max(3, max_len)
            
            for i in range(grid_rows):
                # Column A: Dual on row 0, Focused on row 1, empty after
                if i == 0:
                    ws.write(row + i, 0, f"Dual: {dual_counted:.1f} ({dual_actual:.1f})", stats_fmt)
                elif i == 1:
                    ws.write(row + i, 0, f"Focused: {focused:.1f}", stats_fmt)
                
                # Schedule data in columns B onwards
                for idx, day in enumerate(days):
                    entries = intervals.get(day, [])
                    name_col = 1 + idx * 2
                    time_col = 1 + idx * 2 + 1
                    
                    if i < len(entries):
                        name, start, end = entries[i]
                        start_str = time_slot_starts[start]
                        end_str = time_slot_starts[end]
                        ws.write(row + i, name_col, name, cell_fmt)
                        ws.write(row + i, time_col, _format_time_range(start_str, end_str), cell_fmt)
                    else:
                        # Empty cell with border
                        ws.write(row + i, name_col, "", empty_cell_fmt)
                        ws.write(row + i, time_col, "", empty_cell_fmt)
            
            row += grid_rows

    workbook.close()


def _autosize_columns(writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame):
    worksheet = writer.sheets[sheet_name]
    for idx, column in enumerate(dataframe.columns):
        max_len = max([len(str(column))] + [len(str(cell)) for cell in dataframe[column]])
        worksheet.set_column(idx, idx, max_len + 2)


def _write_minimal_xlsx(output_path: Path, weekly_columns: List[str], weekly_rows: List[List]):
    """Fallback XLSX writer if no Excel engines are installed."""
    sheets = [
        (
            "xl/worksheets/sheet1.xml",
            _create_sheet_xml("Weekly Grid", weekly_columns, weekly_rows),
        )
    ]
    _write_minimal_xlsx_archive(output_path, sheets)


def _create_sheet_xml(name: str, columns: List[str], rows: List[List]) -> str:
    header = "".join(f'<c t="inlineStr"><is><t>{col}</t></is></c>' for col in columns)
    body = ""
    for row in rows:
        body += "<row>" + "".join(f'<c t="inlineStr"><is><t>{cell}</t></is></c>' for cell in row) + "</row>"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData>"
        f"<row>{header}</row>"
        f"{body}"
        "</sheetData>"
        "</worksheet>"
    )


def _write_minimal_xlsx_archive(output_path: Path, sheet_files: List[Tuple[str, str]]):
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        + "".join(
            f'<Override PartName="/{filename}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            for filename, _ in sheet_files
        )
        + "</Types>"
    )

    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )

    sheets_entries = "".join(
        f'<sheet name="Sheet{idx + 1}" sheetId="{idx + 1}" r:id="rId{idx + 1}"/>' for idx, _ in enumerate(sheet_files)
    )
    workbook_rels = [
        f'<Relationship Id="rId{idx + 1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{idx + 1}.xml"/>'
        for idx, _ in enumerate(sheet_files)
    ]

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheets_entries}</sheets>"
        "</workbook>"
    )

    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(workbook_rels)}'
        "</Relationships>"
    )

    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )

    with ZipFile(output_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", rels_xml)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        archive.writestr("xl/styles.xml", styles_xml)
        for filename, xml in sheet_files:
            archive.writestr(filename, xml)
